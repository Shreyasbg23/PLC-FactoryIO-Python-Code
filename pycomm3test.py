from pycomm3 import LogixDriver
import time
import random
import openpyxl


plc1 = LogixDriver('192.168.3.3')
plc1.open()

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active 

i =0

diffsensor6 = 0
diffsensor7 = 0

total_bags   = list( )
temp_bags=list( )
Security_bags = list( )
total_bags_count=0
Security_bags_count=0

plane1 = list( )
plane2 = list( )
plane3 = list( )

output_tags=list(bytearray(64))
bag_main_list=list( )

current_time = 0
previous_time = 0

current_time_1 = 0
Previous_time_1 = 0

current_time_2 = 0
Previous_time_2 = 0

current_time_3 = 0
Previous_time_3 = 0



inputstation = 0        # state to hold which station to send baggage
lineout=0               # check bag goes out
securitylineout=0       #for triggering Security lineout
bufferlineout=0         #for triggering Buffer lineout
leftlineout= 0          #for triggering Exit Left lineout
rightlineout= 0         #for triggering Exit Right lineout
                        # sequence baggage input variable

i=0 
check_trigger1=0        #for diffuse sensor 4
check_trigger2=0        #for diffuse sensor 5
check_trigger3=0        #for diffuse sensor 6
check_trigger4=0        #for diffuse sensor 7
temp_cam0id = 0

#variables for exit sorting
check_trigger5=0 # Plane1
check_trigger6=0 # Plane2
check_trigger7=0 # Plane3
plane_1_temp = 0 # Plane1 temp
plane_2_temp = 0 # Plane2 temp
plane_3_temp = 0 # Plane3 temp



def turn_main_conveyor_on():#output_tags[0 to 30]  main conveyor line
        global output_tags
        for i in range(0,30):
                plc1.write("BOOLOUT"+str(i),True)

def Station_0(state):#output_tags[30 to 32]  bags in line 0
        global output_tags
        for i in range(30,32):
                plc1.write("BOOLOUT"+str(i),state)
                
def Station_3(state):#output_tags[32 to 34]  bags in line 3
        global output_tags
        for i in range(32,34):
                plc1.write("BOOLOUT"+str(i),state)
                
def Station_1(state):#output_tags[35 to 37]  bags in line 1
        global output_tags
        for i in range(35,37):
                plc1.write("BOOLOUT"+str(i),state)
def Station_2(state):#output_tags[37 to 39]  bags in line 2
        global output_tags
        for i in range(37,39):
                plc1.write("BOOLOUT"+str(i),state)

def write_to_excel():
        global plane_name,plane_id
        ws.cell(row=1, column=1, value = 'S.No')
        ws.cell(row=1, column=2, value = 'BAG ID')
        ws.cell(row=1, column=3, value = 'FLIGHT')
        for i in range(2,len(plane_name)+1):
                ws.cell(row=i, column=1, value = i-2)
        for i in range(2,len(plane_name)+2):
                ws.cell(row=i, column=2, value = plane_id[i-2])
        for i in range(2,len(plane_name)+2):
                ws.cell(row=i, column=3, value = plane_name[i-2])
        wb.save('data.xlsx')



def write_array():
        g=0
        if(len(total_bags)>0):
                while(g<len(total_bags)):
                        plc1.write("totalbags["+str(g)+"]",total_bags[g])
                        g=g+1
        g=0
        if(len(Security_bags)>0):
                while(g<len(Security_bags)):
                        plc1.write("securitybags["+str(g)+"]",Security_bags[g])
                        g=g+1
        g=0
        if(len(plane1)>0):
                while(g<len(plane1)):
                        plc1.write("plane1["+str(g)+"]",plane1[g])
                        g=g+1
        g=0
        if(len(plane2)>0):
                while(g<len(plane2)):
                        plc1.write("plane2["+str(g)+"]",plane2[g])
                        g=g+1
        g=0
        if(len(plane3)>0):
                while(g<len(plane3)):
                        plc1.write("plane3["+str(g)+"]",plane3[g])
                        g=g+1






while True:
        plc1.write("CONNECTION_CHECKBIT",True)
        turn_main_conveyor_on()
        
        if(lineout==0):
                inputstation = random.randint(1,3)
                if (inputstation == 0):
                        Station_0(True)
                        Station_1(False)
                        Station_2(False)
                        Station_3(False)
                if (inputstation == 1):
                        Station_0(False)
                        Station_1(True)
                        Station_2(False)
                        Station_3(False)
                if (inputstation == 2):
                        Station_0(False)
                        Station_1(False)
                        Station_2(True)
                        Station_3(False)
                if (inputstation == 3):
                        Station_0(False)
                        Station_1(False)
                        Station_2(False)
                        Station_3(True)
                i=i+1
                lineout=1
        if(plc1.read("BOOLIN4")[1]):
                check_trigger1=1
                temp_cam0id=plc1.read("INTREAD1")[1]
        if(check_trigger1==1 and plc1.read("BOOLIN4")[1]== False):
                total_bags.append(temp_cam0id)
                temp_bags.append(temp_cam0id)
                total_bags_count+=1
                lineout=0
                check_trigger1=0
        #sorting  for security check
        if(plc1.read("BOOLIN11")[1]==True or plc1.read("BOOLIN12")[1]==True ):
                check_trigger2 = 1
        if(check_trigger2==1 and (plc1.read("BOOLIN11")[1]==False or plc1.read("BOOLIN12")[1]==False) ):
                plc1.write("BOOLOUT40",True)
                previous_time = round(time.time()*10)
                securitylineout = 1
                Security_bags_count+=1
                check_trigger2=0
        current_time = round(time.time()*10)
        if(securitylineout == 1 and (current_time-previous_time)>15):
                plc1.write("BOOLOUT40",False)
                #Security_bags.append(temp_bags.pop(Security_bags_count))
                securitylineout = 0

        #bufffer Re entry
        if(plc1.read("BOOLIN14")[1]):
                check_trigger3=1
        if(check_trigger3==1 and plc1.read("BOOLIN14")[1]==False):
                diffsensor6+=1
        if(plc1.read("BOOLIN15")[1]):
                check_trigger4=1
        if(check_trigger4==1 and plc1.read("BOOLIN15")[1]==False):
                diffsensor7+=1
        if(plc1.read("BOOLIN13")[1] ):
                if(diffsensor6-diffsensor7==0):
                        plc1.write("BOOLOUT41",False)
                        Previous_time_1 = round(time.time()*10)
                        bufferlineout=1
                else:
                        plc1.write("BOOLOUT41",True)
        current_time_1 = round(time.time()*10)
        if(bufferlineout == 1 and (current_time_1-Previous_time_1)>15):
                plc1.write("BOOLOUT41",True)
                bufferlineout = 0
#exit sorting Station
        #right plane
        if(plc1.read("INTREAD0")[1]==4):
                plc1.write("BOOLOUT46",True)
                Previous_time_2= round(time.time()*10)
                leftlineout= 1;
                if(plc1.read("INTREAD2")[1] not in plane1 and (plc1.read("INTREAD2")[1] is not 0)):
                        plane1.append(plc1.read("INTREAD2")[1])
        current_time_2 = round(time.time()*10)
        if(leftlineout == 1 and (current_time_2-Previous_time_2)>5):
                plc1.write("BOOLOUT46",False)
                leftlineout = 0
        
        #Left plane       
        if(plc1.read("INTREAD0")[1]==1):
                plc1.write("BOOLOUT47",True)
                Previous_time_3= round(time.time()*10)
                rightlineout= 1
                if(plc1.read("INTREAD2")[1] not in plane3 and (plc1.read("INTREAD2")[1] is not 0)):
                        plane3.append(plc1.read("INTREAD2")[1])
        current_time_3 = round(time.time()*10)
        if(rightlineout == 1 and (current_time_3-Previous_time_3)>5):
                plc1.write("BOOLOUT47",False)
                rightlineout = 0
        
        #mid plane       
        if(plc1.read("INTREAD0")[1]==7):
                if(plc1.read("INTREAD2")[1] not in plane2 and (plc1.read("INTREAD2")[1] is not 0)):
                        plane2.append(plc1.read("INTREAD2")[1])
        write_array()
        plane_name= list( )
        plane_id = list( )
        for i in range(0,len(plane1)) :
            for k in range(0,len(total_bags)):
                if total_bags[k] == plane1[i]:
                    plane_name.append("United")
                    plane_id.append(plane1[i])
        for i in range(0,len(plane2)) :
            for k in range(0,len(total_bags)):
                if total_bags[k] == plane2[i]:
                    plane_name.append("Air Asia") 
                    plane_id.append(plane2[i])
        for i in range(0,len(plane3)):
            for k in range(0,len(total_bags)):
                if total_bags[k] == plane3[i]:
                    plane_name.append("Star") 
                    plane_id.append(plane3[i])
        write_to_excel()
        plc1.write("CONNECTION_CHECKBIT",False)



def deenergize():
        for i in range(0,64):
                plc1.write("BOOLOUT["+str(i)+"]",False)





























def read_tags():
        global input_tags
        for i in range(0,32):
                input_tags[i]=plc1.read("Local:1:I.Data."+str(i))[1]
def write_tags():
        global output_tags
        for i in range(0,16):
                plc1.write("Local:2:O.Data."+str(i),output_tags[i])

def print_tags():
        global input_tags,output_tags
        read_tags()
        for i in range(0,16):
                print("Local:1:I.Data."+str(i),input_tags[i])
        
#read_tags()
#write_tags()
#print_tags()

